"""
create_unified_template.py
Генератор единого Excel-шаблона для всех вкладок приложения.

Листы:
  Гидропост   — ряд наблюдений для статистики/кривых/трендов
  Работа1     — норма годового стока (расчётная река + аналог)
  Работа2     — внутригодовое распределение (помесячные расходы)
  Работа3     — минимальный сток (зимние и летние минимумы)
  Работа4     — максимальный сток (ряды максимумов по периодам осреднения)
  Работа5     — ледовые явления (даты ледостава и вскрытия)
  Работа6     — водный баланс (суточные/посты данные)
  Работа7     — ливневый сток (параметры расчёта: площадь, климатическая зона)
  Работа8     — кривая обеспеченности продолжительности (FDC)
  Работа9     — гидротехнические расчёты (параметры: расход, ширина, уклон)
  Работа10    — экология и базовый сток
  ГТС         — параметры ГТС для классификации

Пустые листы — данные по ним не загружаются.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment


def create_unified_template(path="unified_template.xlsx"):
    """Создаёт Excel-шаблон со всеми листами."""
    wb = Workbook()

    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    info_font = Font(italic=True, color="C00000", size=9)
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    def style_cell(ws, row, col, value, font=None, fill=None):
        cell = ws.cell(row=row, column=col, value=value)
        if font: cell.font = font
        if fill: cell.fill = fill
        cell.border = thin
        cell.alignment = Alignment(horizontal='center')
        return cell

    def write_header_row(ws, row, headers):
        for c, h in enumerate(headers, 1):
            style_cell(ws, row, c, h, font=header_fill, fill=header_fill)
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin
            cell.alignment = Alignment(horizontal='center')

    # ========== ЛИСТ 1: Гидропост ==========
    ws1 = wb.active
    ws1.title = "Гидропост"
    ws1['A1'] = "Данные гидрологического поста для статистической обработки"
    ws1['A1'].font = Font(bold=True, size=12)
    ws1.merge_cells('A1:D1')
    ws1['A2'] = "Заполни данные в жёлтых ячейках или оставь как есть"
    ws1['A2'].font = info_font
    ws1.merge_cells('A2:D2')

    # Пытаемся прочитать test_data_clean.xlsx
    post_data = {}
    post_names = []
    try:
        df_test = pd.read_excel("test_data_clean.xlsx")
        year_col = None
        for c in df_test.columns:
            if str(c).strip().lower() in ['год', 'year']:
                year_col = c
                break
        if year_col:
            post_cols = [c for c in df_test.columns if c != year_col]
            for pc in post_cols:
                series = {}
                for _, row in df_test.iterrows():
                    y = row[year_col]
                    q = row[pc]
                    if pd.notna(y) and pd.notna(q):
                        series[int(y)] = round(float(q), 2)
                if series:
                    post_names.append(f"Пост {pc}")
                    post_data[f"Пост {pc}"] = series
    except (FileNotFoundError, ValueError, KeyError, TypeError):
        pass  # шаблон создаётся с дефолтными данными

    if not post_data:
        post_names = ["Расход Q, м³/с"]
        post_data[post_names[0]] = dict(zip(range(1960, 1980), [
            120, 145, 98, 156, 134, 110, 167, 129, 143, 108,
            152, 118, 136, 147, 102, 158, 125, 139, 151, 106
        ]))

    write_header_row(ws1, 7, ["Год"] + post_names)
    all_years = sorted({y for s in post_data.values() for y in s})
    for i, y in enumerate(all_years, start=8):
        style_cell(ws1, i, 1, y, fill=yellow_fill)
        for col, pname in enumerate(post_names, start=2):
            style_cell(ws1, i, col, post_data[pname].get(y, None), fill=yellow_fill)

    # ========== ЛИСТ 2: Работа1 ==========
    ws2 = wb.create_sheet("Работа1")
    ws2['A1'] = "Работа 1 — Норма годового стока"
    ws2['A1'].font = Font(bold=True, size=12)
    ws2.merge_cells('A1:E1')

    # Расчётная река
    ws2['A3'] = "РАСЧЁТНАЯ РЕКА"
    ws2['A3'].font = Font(bold=True, color="1F4E79")
    ws2['A4'] = "Название:"
    ws2['B4'] = "Бирюса, с. Шиткино"
    ws2['B4'].fill = yellow_fill
    ws2['A5'] = "Площадь F, км²:"
    ws2['B5'] = 31800
    ws2['B5'].fill = yellow_fill

    write_header_row(ws2, 7, ["Год", "Q, м³/с"])
    calc_years = list(range(1944, 1976))
    calc_Q = [361,210,364,250,225,241,287,284,403,274,306,395,243,291,
              253,264,276,288,268,266,243,318,376,308,310,226,226,260,307,372,203,290]
    for i, (y, q) in enumerate(zip(calc_years, calc_Q), start=8):
        style_cell(ws2, i, 1, y, fill=yellow_fill)
        style_cell(ws2, i, 2, q, fill=yellow_fill)

    # Река-аналог
    ws2['A45'] = "РЕКА-АНАЛОГ"
    ws2['A45'].font = Font(bold=True, color="1F4E79")
    ws2['A46'] = "Название:"
    ws2['B46'] = "Бирюса, р.п. Суетиха"
    ws2['B46'].fill = yellow_fill
    ws2['A47'] = "Площадь F, км²:"
    ws2['B47'] = 24700
    ws2['B47'].fill = yellow_fill

    write_header_row(ws2, 49, ["Год", "Q, м³/с"])
    analog_years = list(range(1936, 1976))
    analog_Q = [335,280,276,292,252,339,200,159,329,177,320,229,203,221,262,249,368,
                243,283,367,218,256,223,250,260,269,252,244,215,297,341,280,203,250,369,252,278,323,181,264]
    for i, (y, q) in enumerate(zip(analog_years, analog_Q), start=50):
        style_cell(ws2, i, 1, y, fill=yellow_fill)
        style_cell(ws2, i, 2, q, fill=yellow_fill)

    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 16

    # ========== ЛИСТ 3: Работа2 ==========
    ws3 = wb.create_sheet("Работа2")
    ws3['A1'] = "Работа 2 — Внутригодовое распределение (помесячные расходы)"
    ws3['A1'].font = Font(bold=True, size=12)
    ws3.merge_cells('A1:M1')
    ws3['A2'] = "Год | I | II | III | IV | V | VI | VII | VIII | IX | X | XI | XII"
    ws3['A2'].font = info_font
    ws3.merge_cells('A2:M2')

    write_header_row(ws3, 4, ["Год", "I", "II", "III", "IV", "V", "VI", "VII", "VIII", "IX", "X", "XI", "XII"])

    for yr in range(1965, 1976):
        style_cell(ws3, yr - 1965 + 5, 1, yr, fill=yellow_fill)
        for m in range(1, 13):
            import random
            random.seed(yr * 100 + m)
            style_cell(ws3, yr - 1965 + 5, m + 1,
                       round(random.uniform(5, 60), 1), fill=yellow_fill)

    # ========== ЛИСТ 4: Работа3 ==========
    ws4 = wb.create_sheet("Работа3")
    ws4['A1'] = "Работа 3 — Минимальный сток (30-суточные минимумы)"
    ws4['A1'].font = Font(bold=True, size=12)
    ws4.merge_cells('A1:D1')

    write_header_row(ws4, 3, ["Год", "Зимний минимум", "Летний минимум", "Примечание"])
    for yr in range(1970, 1991):
        import random
        random.seed(yr * 10)
        style_cell(ws4, yr - 1970 + 4, 1, yr, fill=yellow_fill)
        style_cell(ws4, yr - 1970 + 4, 2, round(random.uniform(1.5, 8.0), 2), fill=yellow_fill)
        style_cell(ws4, yr - 1970 + 4, 3, round(random.uniform(3.0, 15.0), 2), fill=yellow_fill)
        style_cell(ws4, yr - 1970 + 4, 4, "—")

    # ========== ЛИСТ 5: Работа4 ==========
    ws4b = wb.create_sheet("Работа4")
    ws4b['A1'] = "Работа 4 — Максимальный сток (ряды максимумов)"
    ws4b['A1'].font = Font(bold=True, size=12)
    ws4b.merge_cells('A1:E1')
    ws4b['A2'] = "Год | Q_max (м³/с) | Период осреднения: 1 сутки | 5 суток | 7 суток"
    ws4b['A2'].font = info_font
    ws4b.merge_cells('A2:E2')

    write_header_row(ws4b, 4, ["Год", "Q_max, м³/с", "Q_5сут, м³/с", "Q_7сут, м³/с", "Примечание"])
    for i, yr in enumerate(range(1970, 1991)):
        import random
        random.seed(yr * 3 + 1)
        style_cell(ws4b, 5 + i, 1, yr, fill=yellow_fill)
        style_cell(ws4b, 5 + i, 2, round(random.uniform(300, 900), 1), fill=yellow_fill)
        style_cell(ws4b, 5 + i, 3, round(random.uniform(250, 750), 1), fill=yellow_fill)
        style_cell(ws4b, 5 + i, 4, round(random.uniform(220, 680), 1), fill=yellow_fill)

    # ========== ЛИСТ 6: Работа5 ==========
    ws5b = wb.create_sheet("Работа5")
    ws5b['A1'] = "Работа 5 — Ледовые явления"
    ws5b['A1'].font = Font(bold=True, size=12)
    ws5b.merge_cells('A1:C1')
    ws5b['A2'] = "Год | Дата ледостава | Дата вскрытия"
    ws5b['A2'].font = info_font
    ws5b.merge_cells('A2:C2')

    write_header_row(ws5b, 4, ["Год", "Дата ледостава", "Дата вскрытия"])
    for i, yr in enumerate(range(1970, 1991)):
        style_cell(ws5b, 5 + i, 1, yr, fill=yellow_fill)
        style_cell(ws5b, 5 + i, 2, f"{yr}-11-20", fill=yellow_fill)
        style_cell(ws5b, 5 + i, 3, f"{yr + 1}-04-10", fill=yellow_fill)

    # ========== ЛИСТ 7: Работа6 ==========
    ws6b = wb.create_sheet("Работа6")
    ws6b['A1'] = "Работа 6 — Водный баланс"
    ws6b['A1'].font = Font(bold=True, size=12)
    ws6b.merge_cells('A1:C1')

    write_header_row(ws6b, 3, ["Год", "Осадки, мм", "Сток, мм", "Испарение, мм"])
    for i, yr in enumerate(range(1980, 2000)):
        import random
        random.seed(yr * 7)
        style_cell(ws6b, 4 + i, 1, yr, fill=yellow_fill)
        style_cell(ws6b, 4 + i, 2, round(random.uniform(500, 700), 1), fill=yellow_fill)
        style_cell(ws6b, 4 + i, 3, round(random.uniform(150, 300), 1), fill=yellow_fill)
        style_cell(ws6b, 4 + i, 4, round(random.uniform(250, 400), 1), fill=yellow_fill)

    # ========== ЛИСТ 8: Работа7 ==========
    ws7b = wb.create_sheet("Работа7")
    ws7b['A1'] = "Работа 7 — Ливневый сток (параметры расчёта)"
    ws7b['A1'].font = Font(bold=True, size=12)
    ws7b.merge_cells('A1:B1')

    ws7b['A3'] = "Площадь бассейна F, км²:"
    ws7b['B3'] = 25
    ws7b['B3'].fill = yellow_fill
    ws7b['A4'] = "Климатическая зона (zone_1 ... zone_7):"
    ws7b['B4'] = "zone_3"
    ws7b['B4'].fill = yellow_fill
    ws7b['A5'] = "Обеспеченность T, лет:"
    ws7b['B5'] = 10
    ws7b['B5'].fill = yellow_fill
    ws7b['A6'] = "Время концентрации t, мин:"
    ws7b['B6'] = 60
    ws7b['B6'].fill = yellow_fill
    ws7b['A7'] = "Коэфф. стока α:"
    ws7b['B7'] = 0.70
    ws7b['B7'].fill = yellow_fill

    ws7b.column_dimensions['A'].width = 35
    ws7b.column_dimensions['B'].width = 15

    # ========== ЛИСТ 9: Работа8 (FDC) ==========
    ws8b = wb.create_sheet("Работа8")
    ws8b['A1'] = "Работа 8 — Кривая обеспеченности продолжительности (FDC)"
    ws8b['A1'].font = Font(bold=True, size=12)
    ws8b.merge_cells('A1:B1')

    write_header_row(ws8b, 3, ["Год", "Расход Q, м³/с"])
    for i in range(30):
        import random
        random.seed(1000 + i)
        style_cell(ws8b, 4 + i, 1, 1980 + i, fill=yellow_fill)
        style_cell(ws8b, 4 + i, 2, round(random.uniform(50, 300), 1), fill=yellow_fill)

    # ========== ЛИСТ 10: Работа9 ==========
    ws9b = wb.create_sheet("Работа9")
    ws9b['A1'] = "Работа 9 — Гидротехнические расчёты (параметры)"
    ws9b['A1'].font = Font(bold=True, size=12)
    ws9b.merge_cells('A1:B1')

    ws9b['A3'] = "Расход паводка Q, м³/с:"
    ws9b['B3'] = 500
    ws9b['B3'].fill = yellow_fill
    ws9b['A4'] = "Ширина B, м:"
    ws9b['B4'] = 45
    ws9b['B4'].fill = yellow_fill
    ws9b['A5'] = "Уклон I, м/м (или ‰):"
    ws9b['B5'] = 0.002
    ws9b['B5'].fill = yellow_fill

    ws9b.column_dimensions['A'].width = 35
    ws9b.column_dimensions['B'].width = 15

    # ========== ЛИСТ 11: Работа10 ==========
    ws10b = wb.create_sheet("Работа10")
    ws10b['A1'] = "Работа 10 — Экология и базовый сток"
    ws10b['A1'].font = Font(bold=True, size=12)
    ws10b.merge_cells('A1:B1')

    write_header_row(ws10b, 3, ["Год", "Базовый сток, м³/с", "Qэкологический, м³/с"])
    for i, yr in enumerate(range(1980, 2000)):
        import random
        random.seed(yr * 5 + 2)
        style_cell(ws10b, 4 + i, 1, yr, fill=yellow_fill)
        style_cell(ws10b, 4 + i, 2, round(random.uniform(20, 90), 1), fill=yellow_fill)
        style_cell(ws10b, 4 + i, 3, round(random.uniform(25, 100), 1), fill=yellow_fill)

    # ========== ЛИСТ 12: ГТС ==========
    ws5 = wb.create_sheet("ГТС")
    ws5['A1'] = "Параметры ГТС для классификации"
    ws5['A1'].font = Font(bold=True, size=12)

    ws5['A3'] = "Высота плотины, м:"
    ws5['B3'] = 35
    ws5['B3'].fill = yellow_fill

    ws5['A4'] = "Объём водохранилища, млн.м³:"
    ws5['B4'] = 200
    ws5['B4'].fill = yellow_fill

    ws5.column_dimensions['A'].width = 35
    ws5.column_dimensions['B'].width = 15

    # ========== Сохранение ==========
    wb.save(path)
    print(f"✅ Единый шаблон создан: {path}")
    print(f"   Листы: {wb.sheetnames}")
    return path


if __name__ == "__main__":
    create_unified_template()
