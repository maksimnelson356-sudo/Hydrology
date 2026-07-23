"""
core/hydrorash/intra_annual.py
Модуль внутригодового распределения стока

Перенесено из HydroRash с адаптацией под hydrolib.
"""

import pandas as pd
import numpy as np
from typing import Dict, List, Optional
from .hydrological_periods import HydrologicalPeriods


def calculate_water_year_sums(
    monthly_df: pd.DataFrame,
    periods: Optional[HydrologicalPeriods] = None,
    month_columns: Optional[List[str]] = None
) -> pd.DataFrame:
    """Расчёт сумм стока по периодам водного года."""
    if periods is None:
        periods = HydrologicalPeriods()

    df = monthly_df.copy()

    if month_columns is None:
        month_columns = []
        numeric_cols = [col for col in df.columns if isinstance(col, (int, float)) and 1 <= int(col) <= 12]
        if numeric_cols:
            month_columns = [int(c) for c in numeric_cols]

        if not month_columns:
            month_map = {
                "I": 1, "II": 2, "III": 3, "IV": 4, "V": 5, "VI": 6,
                "VII": 7, "VIII": 8, "IX": 9, "X": 10, "XI": 11, "XII": 12,
                "ЯНВ": 1, "ФЕВ": 2, "МАР": 3, "АПР": 4, "МАЙ": 5, "ИЮН": 6,
                "ИЮЛ": 7, "АВГ": 8, "СЕН": 9, "ОКТ": 10, "НОЯ": 11, "ДЕК": 12
            }
            for col in df.columns:
                col_str = str(col).strip().upper().replace(".", "")
                if col_str in month_map:
                    month_columns.append(month_map[col_str])

    if not month_columns:
        raise ValueError("Не удалось определить столбцы с месяцами")

    month_columns = sorted(set([int(m) for m in month_columns]))

    rename_dict = {}
    for col in df.columns:
        col_str = str(col).strip().upper()
        if col_str in ["I","II","III","IV","V","VI","VII","VIII","IX","X","XI","XII"]:
            rename_dict[col] = {"I":1,"II":2,"III":3,"IV":4,"V":5,"VI":6,
                                "VII":7,"VIII":8,"IX":9,"X":10,"XI":11,"XII":12}[col_str]

    if rename_dict:
        df = df.rename(columns=rename_dict)

    result = pd.DataFrame(index=df.index)
    result["год"] = df.index

    all_months = periods.non_limiting_months + periods.limiting_months
    nlp_months = [m for m in periods.non_limiting_months if m in df.columns]
    lp_months  = [m for m in periods.limiting_months if m in df.columns]
    ls_months  = [m for m in periods.limiting_season_months if m in df.columns] if periods.limiting_season_months else lp_months

    result["сумма_год"] = df[[m for m in all_months if m in df.columns]].sum(axis=1)
    result["сумма_НЛП"] = df[nlp_months].sum(axis=1) if nlp_months else 0
    result["сумма_ЛП"]  = df[lp_months].sum(axis=1) if lp_months else 0
    result["сумма_ЛС"]  = df[ls_months].sum(axis=1) if ls_months else result["сумма_ЛП"]

    return result.reset_index(drop=True)


def compute_intra_annual_stats(
    sums_df: pd.DataFrame,
    columns: Optional[List[str]] = None
) -> Dict[str, Dict[str, float]]:
    from .utils import compute_basic_stats

    if columns is None:
        columns = ["сумма_год", "сумма_НЛП", "сумма_ЛП", "сумма_ЛС"]

    available_cols = [col for col in columns if col in sums_df.columns]
    if not available_cols:
        raise ValueError(f"Нет колонок: {columns}")

    stats = {}
    for col in available_cols:
        series = sums_df[col].dropna()
        if len(series) < 3:
            stats[col] = {"mean": None, "Cv": None, "Cs": None, "epsilon": None, "n": len(series)}
            continue
        basic = compute_basic_stats(series, use_normative_Cs=True)
        stats[col] = {
            "mean": basic["mean"],
            "Cv": basic["Cv"],
            "Cs": basic["Cs"],
            "Cs/Cv": basic["Cs/Cv"],
            "epsilon": basic["epsilon"],
            "n": basic["n"]
        }
    return stats


def select_model_year(
    sums_df: pd.DataFrame,
    target_P: float = 90.0,
    by: str = "сумма_ЛП"
) -> Dict:
    from .utils import kritsky_menkel_quantiles

    if by not in sums_df.columns:
        raise ValueError(f"Колонка '{by}' не найдена")

    stats = compute_intra_annual_stats(sums_df, columns=[by])
    mean_val = stats[by]["mean"]
    Cv = stats[by]["Cv"]

    km = kritsky_menkel_quantiles(mean_val, Cv, stats[by].get("Cs/Cv", 2.0), [target_P])
    target_sum = km.iloc[0]["Q_p"]

    sums_df = sums_df.copy()
    sums_df["delta"] = abs(sums_df[by] - target_sum)
    best = sums_df.loc[sums_df["delta"].idxmin()]

    return {
        "год": best.get("год", best.name),
        "сумма_ЛП": best.get("сумма_ЛП"),
        "сумма_ЛС": best.get("сумма_ЛС"),
        "delta": best["delta"],
        "target_sum": target_sum,
        "target_P": target_P
    }


def distribute_discharge(
    annual_sum_P: float,
    model_year_row: pd.Series,
    periods: Optional[HydrologicalPeriods] = None
) -> pd.DataFrame:
    """Распределение годового стока по месяцам по году-модели."""
    if periods is None:
        periods = HydrologicalPeriods()

    month_cols = periods.non_limiting_months + periods.limiting_months
    model_sum = model_year_row[month_cols].sum()

    if model_sum == 0:
        raise ValueError("Сумма в году-модели равна нулю")

    percentages = (model_year_row[month_cols] / model_sum * 100).round(2)
    distributed = (percentages / 100 * annual_sum_P).round(2)

    result = pd.DataFrame({
        "Месяц": month_cols,
        "Расход, м³/с": distributed.values,
        "Доля, %": percentages.values
    })
    return result


def generate_intra_annual_report(
    sums_df: pd.DataFrame,
    stats: dict,
    model_year: Optional[dict] = None,
    periods: Optional[HydrologicalPeriods] = None,
    output_path: str = None
) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Border, Side
    from pathlib import Path

    if periods is None:
        periods = HydrologicalPeriods()

    output_dir = Path(output_path).parent if output_path else Path("reports")
    output_dir.mkdir(parents=True, exist_ok=True)
    final_path = output_dir / "Отчёт_Работа2.xlsx"

    wb = Workbook()
    title_font = Font(bold=True, size=13, color="1F4E79")

    ws = wb.active
    ws.title = "Основные результаты"
    ws['A1'] = "ВНУТРИГОДОВОЕ РАСПРЕДЕЛЕНИЕ СТОКА"
    ws['A1'].font = title_font
    ws['A2'] = "Расчёт согласно СП 33-101-2003, СП 529.1325800.2023"

    row = 4
    for name, s in stats.items():
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=round(s.get('mean', 0), 2) if s.get('mean') else "—")
        ws.cell(row=row, column=3, value=round(s.get('Cv', 0), 4) if s.get('Cv') else "—")
        row += 1

    wb.save(final_path)
    return str(final_path)
