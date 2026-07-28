"""
core/stats/data_loader.py
Поддержка нескольких постов + выбор поста
"""

import pandas as pd


def load_hydrological_data(filepath):
    """
    Загружает файл с несколькими гидропостами.
    Возвращает:
    - df_raw: исходный датафрейм
    - available_posts: список доступных постов
    """
    df_raw = pd.read_excel(filepath)
    df_raw.columns = [str(col).strip() for col in df_raw.columns]
    
    # Ищем столбец с годами
    year_col = None
    for col in df_raw.columns:
        if col.lower() in ['год', 'year', 'years', 'дата', 'date']:
            year_col = col
            break
    if year_col is None:
        year_col = df_raw.columns[0]
    
    # Находим все числовые столбцы (посты)
    available_posts = []
    for col in df_raw.columns:
        if col != year_col:
            try:
                pd.to_numeric(df_raw[col], errors='raise')
                available_posts.append(col)
            except (ValueError, TypeError):
                pass  # столбец не является числовым — пропускаем
    
    return df_raw, year_col, available_posts


def get_series_by_post(df_raw, year_col, post_name):
    """
    Возвращает DataFrame с двумя колонками: year и value для выбранного поста
    """
    df = pd.DataFrame()
    df['year'] = pd.to_numeric(df_raw[year_col], errors='coerce')
    df['value'] = pd.to_numeric(df_raw[post_name], errors='coerce')
    
    df = df.dropna(subset=['value']).reset_index(drop=True)
    df.attrs['post'] = post_name
    
    return df


def get_basic_stats(df):
    values = df['value'].dropna()

    return {
        'Пост': df.attrs.get('post', '-'),
        'Количество значений': len(values),
        'Среднее (Qср)': round(values.mean(), 2),
        'Медиана': round(values.median(), 2),
        'Минимум': round(values.min(), 2),
        'Максимум': round(values.max(), 2),
        'Ст. отклонение': round(values.std(), 2),
        'Cv': round(values.std() / values.mean(), 3) if values.mean() != 0 else 0,
        'Cs': round(values.skew(), 3),
        'Пропусков': int(df['value'].isna().sum())
    }


def parse_hydro_data(xlsx_path: str) -> dict:
    """
    Парсит Excel с гидрологическими данными (из Statistica / РГГМУ).

    Поддерживает два формата:
    1. Листовой — каждый лист = ряд (B1: река, B2: пост, 5+ строка: год/расход)
    2. Универсальный — одна таблица с колонками: река, пост, год, расход

    Returns:
        dict: {key: {'river', 'gauge', 'df': DataFrame, 'n', 'period'}}
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    data = {}

    # Определяем формат
    ws0 = wb[wb.sheetnames[0]]
    is_universal = False
    try:
        if ws0.max_row >= 5:
            row5 = list(ws0.iter_rows(min_row=5, max_row=5, values_only=True))[0]
            non_empty = sum(1 for c in row5 if c is not None)
            val_a4 = ws0['A4'].value
            if non_empty >= 4 and val_a4 and str(val_a4).lower() in ['river', 'река']:
                is_universal = True
    except (ValueError, TypeError, AttributeError):
        pass  # не удалось определить формат — пробуем как листовой

    if is_universal:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows(min_row=5, values_only=True):
                if len(row) >= 4 and row[0] and row[1] and row[2] is not None and row[3] is not None:
                    try:
                        river, gauge = str(row[0]).strip(), str(row[1]).strip()
                        year, Q = int(row[2]), float(row[3])
                        key = f"{river}_{gauge}"
                        if key not in data:
                            data[key] = {'river': river, 'gauge': gauge,
                                         'df': __import__('pandas').DataFrame(columns=['year', 'Q'])}
                        data[key]['df'] = __import__('pandas').concat([
                            data[key]['df'],
                            __import__('pandas').DataFrame({'year': [year], 'Q': [Q]})
                        ], ignore_index=True)
                    except (ValueError, TypeError):
                        continue
    else:
        import pandas as pd
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            river = str(ws['B1'].value).strip() if ws['B1'].value else None
            gauge = str(ws['B2'].value).strip() if ws['B2'].value else None
            if not river or not gauge:
                continue
            years, Qs = [], []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if len(row) >= 2 and row[0] is not None and row[1] is not None:
                    try:
                        years.append(int(row[0]))
                        Qs.append(float(row[1]))
                    except (ValueError, TypeError):
                        continue
            if len(years) >= 3:
                key = f"{river}_{gauge}"
                data[key] = {
                    'river': river, 'gauge': gauge,
                    'df': pd.DataFrame({'year': years, 'Q': Qs})
                }

    wb.close()

    for key in list(data.keys()):
        df = data[key]['df']
        df = df.drop_duplicates(subset='year').sort_values('year').reset_index(drop=True)
        if len(df) >= 3:
            data[key]['df'] = df
            data[key]['n'] = len(df)
            data[key]['period'] = f"{int(df['year'].min())}-{int(df['year'].max())}"
        else:
            del data[key]

    return data