"""
core/stats/report.py
Формирование и сохранение отчётов
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict


def save_report_to_excel(stats: Dict, 
                         frequency_curve: pd.DataFrame,
                         filepath: str = "report.xlsx"):
    """
    Сохранение красивого отчёта в Excel.
    """
    wb = Workbook()
    
    # === Лист "Статистика" ===
    ws1 = wb.active
    ws1.title = "Статистика"
    
    ws1['A1'] = "СТАТИСТИЧЕСКАЯ ОБРАБОТКА РЯДА"
    ws1['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws1.merge_cells('A1:C1')
    
    ws1['A3'] = "Базовая статистика"
    ws1['A3'].font = Font(bold=True, size=12)
    
    row = 5
    for key, value in stats.items():
        ws1[f'A{row}'] = key
        ws1[f'B{row}'] = value
        ws1[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # === Лист "Кривая обеспеченности" ===
    ws2 = wb.create_sheet("Кривая обеспеченности")
    
    ws2['A1'] = "КРИВАЯ ОБЕСПЕЧЕННОСТИ"
    ws2['A1'].font = Font(bold=True, size=16, color="1F4E79")
    ws2.merge_cells('A1:C1')
    
    # Заголовок таблицы
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for r_idx, row in enumerate(dataframe_to_rows(frequency_curve, index=False, header=True), 3):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 3:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
    
    # Ширина колонок
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 15
    
    wb.save(filepath)
    print(f"Отчёт сохранён: {filepath}")


if __name__ == "__main__":
    print("Модуль core/stats/report.py загружен успешно.")